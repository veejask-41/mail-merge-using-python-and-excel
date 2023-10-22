import openpyxl
from os import path


def load_workbook(wb_path):
    if (path.exists(wb_path)):
        return openpyxl.load_workbook(wb_path)
    else:
        return openpyxl.Workbook()


wb_path = "data.xlsx"

wb = load_workbook(wb_path)

sheet = wb["Sheet1"]

for id, row in enumerate(sheet.values):
    print(id, row)


wb.save(wb_path)
