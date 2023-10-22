from dotenv import load_dotenv
import os
from email.message import EmailMessage
import ssl
import smtplib
import openpyxl
from os import path

load_dotenv()

email_sender = os.getenv('EMAIL_SENDER')
email_password = os.getenv('EMAIL_APP_PASSWORD_FOR_PYTHON')


def load_workbook(wb_path):
    if (path.exists(wb_path)):
        return openpyxl.load_workbook(wb_path)
    else:
        return openpyxl.Workbook()


wb_path = "data.xlsx"

wb = load_workbook(wb_path)

sheet = wb["Sheet1"]

for id, row in enumerate(sheet.values):
    if (id == 0):
        pass
    else:
        email_subject = "test email to {name}".format(name=row[0])
        email_body = "Hi {name}".format(name=row[0])
        email_recipient = row[1]

        em = EmailMessage()
        em['Subject'] = email_subject
        em['From'] = email_sender
        em['To'] = email_recipient
        em.set_content(email_body)

        ssl_context = ssl.create_default_context()

        with smtplib.SMTP_SSL("smtp.gmail.com", 465, context=ssl_context) as server:
            server.login(email_sender, email_password)
            server.send_message(em)
