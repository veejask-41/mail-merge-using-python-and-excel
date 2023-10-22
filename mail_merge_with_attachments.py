from dotenv import load_dotenv
import os
from email.message import EmailMessage
import ssl
import smtplib
import openpyxl
from os import path
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

load_dotenv()

email_sender = os.getenv('EMAIL_SENDER')
email_password = os.getenv('EMAIL_APP_PASSWORD_FOR_PYTHON')


def load_workbook(wb_path):
    if (path.exists(wb_path)):
        return openpyxl.load_workbook(wb_path)
    else:
        return openpyxl.Workbook()


wb_path = "data.xlsx"

wb = load_workbook(wb_path)

sheet = wb["Sheet1"]


server = smtplib.SMTP('smtp.gmail.com', 587)
server.starttls()
server.login(email_sender, email_password)

for id, row in enumerate(sheet.values):
    if (id == 0):
        pass
    else:
        msg = MIMEMultipart()

        msg['From'] = email_sender
        msg['To'] = row[1]
        msg['Subject'] = "Hey {name}".format(name=row[0])

        body = "Vanakkam {name} !!!".format(name=row[0])

        msg.attach(MIMEText(body, 'plain'))

        filename = row[2]+".pdf"
        attachment = open("./attachments/"+filename, "rb")

        part = MIMEBase('application', 'octet-stream')
        part.set_payload((attachment).read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition',
                        "attachment; filename= %s" % filename)

        msg.attach(part)

        text = msg.as_string()
        server.sendmail(email_sender, row[1], text)


server.quit()
wb.close()